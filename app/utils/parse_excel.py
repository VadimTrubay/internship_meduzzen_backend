from openpyxl import load_workbook

from typing import List
from app.exept.custom_exceptions import BadRequest
from app.schemas.quizzes import QuizSchema, QuestionSchema


def parse_excel(file_path: str) -> List[QuizSchema]:
    required_columns = [
        "name",
        "description",
        "frequency_days",
        "question_text",
        "correct_answer",
        "answer_options",
    ]

    workbook = load_workbook(file_path)
    sheet = workbook.active

    headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}

    for column in required_columns:
        if column not in headers:
            raise BadRequest(f"Missing required column: {column}")

    quizzes = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        quiz_name = str(row[headers["name"]])
        description = str(row[headers["description"]])
        frequency_days = row[headers["frequency_days"]]
        question_text = str(row[headers["question_text"]])
        correct_answer = (
            str(row[headers["correct_answer"]])
            if row[headers["correct_answer"]]
            else None
        )
        answer_options = (
            row[headers["answer_options"]].split(",")
            if row[headers["answer_options"]]
            else []
        )

        if quiz_name not in quizzes:
            quizzes[quiz_name] = {
                "name": quiz_name,
                "description": description,
                "frequency_days": int(frequency_days),
                "questions": {},
            }

        questions = quizzes[quiz_name]["questions"]
        question = questions.get(question_text)

        if question is None:
            question = {
                "question_text": question_text,
                "correct_answer": [correct_answer] if correct_answer else [],
                "answer_options": answer_options,
            }
            questions[question_text] = question
        else:
            if correct_answer and correct_answer not in question["correct_answer"]:
                question["correct_answer"].append(correct_answer)
            for option in answer_options:
                if option not in question["answer_options"]:
                    question["answer_options"].append(option)

    quizzes_list = [
        QuizSchema(
            name=quiz["name"],
            description=quiz["description"],
            frequency_days=quiz["frequency_days"],
            questions=[
                QuestionSchema(
                    question_text=q["question_text"],
                    correct_answer=[str(ans) for ans in q["correct_answer"]],
                    answer_options=[
                        str(option).strip() for option in q["answer_options"]
                    ],
                )
                for q in quiz["questions"].values()
            ],
        )
        for quiz in quizzes.values()
    ]

    return quizzes_list
